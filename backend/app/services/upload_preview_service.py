from datetime import datetime, time as time_type
from io import BytesIO
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.transaction import Transaction
from app.parsers.decrypt import open_excel_bytes
from app.parsers.snapshots import SnapshotParseResult, parse_snapshots
from app.parsers.transactions import TransactionRow, parse_transactions
from app.services import transaction_source_identity as source_identity
from app.services import transaction_source_lifecycle_service as lifecycle_service
from app.services import upload_preview_change_builder as preview_builder
from app.services import upload_preview_matching as preview_matching
from app.services import upload_preview_models as preview_models


class InvalidUploadWorkbookError(ValueError):
    pass


async def preview_transaction_upload_workbook(
    db_session: AsyncSession,
    file_bytes: bytes,
    excel_password: str | None = None,
) -> preview_models.UploadTransactionPreviewResult:
    preview_plan = await build_transaction_upload_preview_plan_workbook(
        db_session=db_session,
        file_bytes=file_bytes,
        excel_password=excel_password,
    )
    return preview_plan.preview


async def build_transaction_upload_preview_plan_workbook(
    db_session: AsyncSession,
    file_bytes: bytes,
    excel_password: str | None = None,
) -> preview_models.UploadTransactionPreviewPlan:
    parsed_rows, _ = parse_upload_workbook_contents(
        file_bytes=file_bytes,
        excel_password=excel_password,
    )
    return await build_transaction_upload_preview_plan_from_rows(
        db_session=db_session,
        parsed_rows=parsed_rows,
    )


def parse_upload_workbook_contents(
    *,
    file_bytes: bytes,
    excel_password: str | None = None,
) -> tuple[list[TransactionRow], SnapshotParseResult]:
    try:
        workbook_buffer = open_excel_bytes(file_bytes, password=excel_password)
        workbook = load_workbook(BytesIO(workbook_buffer.read()), data_only=True)
        return parse_transactions(workbook), parse_snapshots(workbook)
    except (
        BadZipFile,
        InvalidFileException,
        KeyError,
        TypeError,
        ValueError,
    ) as exc:
        raise InvalidUploadWorkbookError(
            f"Workbook is missing required BankSalad sheets or sections: {exc}"
        ) from exc


async def preview_transaction_upload_from_rows(
    db_session: AsyncSession,
    parsed_rows: list[TransactionRow],
) -> preview_models.UploadTransactionPreviewResult:
    preview_plan = await build_transaction_upload_preview_plan_from_rows(
        db_session=db_session,
        parsed_rows=parsed_rows,
    )
    return preview_plan.preview


async def build_transaction_upload_preview_plan_from_rows(
    db_session: AsyncSession,
    parsed_rows: list[TransactionRow],
) -> preview_models.UploadTransactionPreviewPlan:
    if not parsed_rows:
        empty_changes: tuple[preview_models.UploadPreviewChangeData, ...] = ()
        return preview_models.UploadTransactionPreviewPlan(
            preview=preview_models.UploadTransactionPreviewResult(
                parsed_transaction_count=0,
                safe_change_count=0,
                review_required_count=0,
                change_type_counts=preview_builder.build_change_counts(empty_changes),
                safe_changes=empty_changes,
                review_required_changes=empty_changes,
            ),
            entries=(),
        )

    window_start, window_end = lifecycle_service.transaction_window_bounds(parsed_rows)
    existing_rows = await lifecycle_service.get_imported_transactions_in_window(
        db_session,
        datetime.combine(window_start.date(), time_type.min),
        datetime.combine(window_end.date(), time_type.max),
    )
    safe_entries, review_required_entries = _build_preview_entries(
        parsed_rows,
        existing_rows,
    )
    safe_changes = tuple(entry.change for entry in safe_entries)
    review_required_changes = tuple(entry.change for entry in review_required_entries)
    all_changes = tuple((*safe_changes, *review_required_changes))
    return preview_models.UploadTransactionPreviewPlan(
        preview=preview_models.UploadTransactionPreviewResult(
            parsed_transaction_count=len(parsed_rows),
            safe_change_count=len(safe_changes),
            review_required_count=len(review_required_changes),
            change_type_counts=preview_builder.build_change_counts(all_changes),
            safe_changes=safe_changes,
            review_required_changes=review_required_changes,
        ),
        entries=tuple((*safe_entries, *review_required_entries)),
    )


def _build_preview_entries(
    parsed_rows: list[TransactionRow],
    existing_rows: list[Transaction],
) -> tuple[
    list[preview_models.UploadPreviewPlanEntry],
    list[preview_models.UploadPreviewPlanEntry],
]:
    exact_hash_buckets = preview_matching.build_exact_hash_buckets(existing_rows)
    fallback_buckets = preview_matching.build_fallback_buckets(existing_rows)
    replacement_buckets = preview_matching.build_replacement_buckets(existing_rows)
    reserved_existing_ids: set[int] = set()
    seen_row_hashes: dict[str, int] = {}
    safe_changes: list[preview_models.UploadPreviewPlanEntry] = []
    review_required_changes: list[preview_models.UploadPreviewPlanEntry] = []

    for row in parsed_rows:
        preview_change = _classify_preview_row(
            row,
            exact_hash_buckets,
            fallback_buckets,
            replacement_buckets,
            reserved_existing_ids,
            seen_row_hashes,
        )
        preview_entry = preview_models.build_preview_plan_entry(preview_change, row)
        match preview_change.change_type:
            case "new" | "unchanged" | "source_fields_changed" | "time_shifted":
                safe_changes.append(preview_entry)
            case "possible_replacement" | "possible_duplicate" | "ambiguous":
                review_required_changes.append(preview_entry)
            case "missing_from_latest_export":
                raise AssertionError("incoming rows cannot become missing")
            case unreachable:
                raise AssertionError(f"unsupported change type: {unreachable}")

    for transaction in existing_rows:
        if transaction.id in reserved_existing_ids:
            continue
        missing_change = preview_builder.build_change(
            change_type="missing_from_latest_export",
            review_required=False,
            reason="An imported transaction in the latest preview window is missing from this workbook.",
            row_hash=(
                transaction.source_row_hash
                or source_identity.source_row_hash_from_transaction(transaction)
            ),
            existing=transaction,
            incoming=None,
            candidate_ids=(transaction.id,),
        )
        safe_changes.append(
            preview_models.build_preview_plan_entry(
                missing_change,
                None,
            )
        )

    return safe_changes, review_required_changes


def _classify_preview_row(
    row: TransactionRow,
    exact_hash_buckets: dict[str, list[Transaction]],
    fallback_buckets: dict[source_identity.FallbackSignature, list[Transaction]],
    replacement_buckets: dict[preview_matching.ReplacementSignature, list[Transaction]],
    reserved_existing_ids: set[int],
    seen_row_hashes: dict[str, int],
) -> preview_models.UploadPreviewChangeData:
    row_hash = source_identity.source_row_hash_from_row(row)
    seen_row_hashes[row_hash] = seen_row_hashes.get(row_hash, 0) + 1
    exact_candidates = preview_matching.available_candidates(
        exact_hash_buckets.get(row_hash, []),
        reserved_existing_ids,
    )
    if seen_row_hashes[row_hash] > 1:
        return preview_builder.build_duplicate_change(
            row,
            row_hash,
            exact_candidates,
            "The preview file contains repeated source rows with the same source hash.",
        )
    if len(exact_candidates) > 1:
        reserved_existing_ids.update(candidate.id for candidate in exact_candidates)
        return preview_builder.build_duplicate_change(
            row,
            row_hash,
            exact_candidates,
            "Multiple imported transactions already share the same source hash.",
        )
    if len(exact_candidates) == 1:
        exact_match = exact_candidates[0]
        reserved_existing_ids.add(exact_match.id)
        return preview_builder.build_change(
            change_type="unchanged",
            review_required=False,
            reason="Exact source hash match with an existing imported transaction.",
            row_hash=row_hash,
            existing=exact_match,
            incoming=row,
            candidate_ids=(exact_match.id,),
        )

    fallback_candidates = preview_matching.available_fallback_candidates(
        fallback_buckets,
        row,
        reserved_existing_ids,
    )
    if len(fallback_candidates) > 1:
        reserved_existing_ids.update(candidate.id for candidate in fallback_candidates)
        return preview_builder.build_ambiguous_change(
            row,
            row_hash,
            fallback_candidates,
        )
    if len(fallback_candidates) == 1:
        fallback_match = fallback_candidates[0]
        reserved_existing_ids.add(fallback_match.id)
        return preview_builder.build_matched_safe_change(
            fallback_match,
            row,
            row_hash,
        )

    replacement_candidates = preview_matching.available_candidates(
        replacement_buckets.get(
            preview_matching.replacement_signature_from_row(row),
            [],
        ),
        reserved_existing_ids,
    )
    if len(replacement_candidates) > 1:
        reserved_existing_ids.update(
            candidate.id for candidate in replacement_candidates
        )
        return preview_builder.build_ambiguous_change(
            row,
            row_hash,
            replacement_candidates,
        )
    if len(replacement_candidates) == 1:
        replacement = replacement_candidates[0]
        reserved_existing_ids.add(replacement.id)
        return preview_builder.build_change(
            change_type="possible_replacement",
            review_required=True,
            reason="A non-exact imported transaction could be a replacement for this source row.",
            row_hash=row_hash,
            existing=replacement,
            incoming=row,
            candidate_ids=(replacement.id,),
        )

    return preview_builder.build_change(
        change_type="new",
        review_required=False,
        reason="No imported transaction matched the incoming source row.",
        row_hash=row_hash,
        existing=None,
        incoming=row,
        candidate_ids=(),
    )
