from io import BytesIO

from openpyxl import load_workbook

from app.parsers.transactions import parse_transactions


def test_parse_transactions_extracts_normalized_rows(
    sample_workbook_bytes: bytes,
) -> None:
    workbook = load_workbook(BytesIO(sample_workbook_bytes), data_only=True)

    parsed = parse_transactions(workbook)

    assert len(parsed) == 2357
    assert parsed[0] == {
        "date": workbook["가계부 내역"]["A2"].value.date(),
        "time": workbook["가계부 내역"]["B2"].value,
        "type": "지출",
        "category_major": "카페",
        "category_minor": "미분류",
        "description": "왓커피신시가지점",
        "merchant": "왓커피신시가지점",
        "amount": -4500,
        "currency": "KRW",
        "payment_method": "American Express Gold Card Edition2",
        "memo": None,
    }


def test_parse_transactions_preserves_positive_expense_amounts(
    sample_workbook_bytes: bytes,
) -> None:
    workbook = load_workbook(BytesIO(sample_workbook_bytes), data_only=True)

    parsed = parse_transactions(workbook)

    positive_expense = next(row for row in parsed if row["type"] == "지출" and row["amount"] > 0)

    assert positive_expense["amount"] == 23200
    assert positive_expense["description"] == "쿠팡이츠"
