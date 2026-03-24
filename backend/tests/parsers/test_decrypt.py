from io import BytesIO

from openpyxl import load_workbook

from app.parsers.decrypt import open_excel_bytes


def test_open_excel_bytes_passes_through_non_encrypted_workbook(
    sample_workbook_bytes: bytes,
) -> None:
    buffer = open_excel_bytes(sample_workbook_bytes, password=None)

    workbook = load_workbook(BytesIO(buffer.read()), data_only=True)

    assert workbook.sheetnames == ["뱅샐현황", "가계부 내역"]
