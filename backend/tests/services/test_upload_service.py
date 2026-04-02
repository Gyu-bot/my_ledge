from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.asset_snapshot import AssetSnapshot
from app.models.investment import Investment
from app.models.loan import Loan
from app.models.transaction import Transaction
from app.models.upload_log import UploadLog
from app.parsers.transactions import parse_transactions
from app.services.upload_service import import_transactions_from_workbook


async def test_import_transactions_inserts_all_rows_on_first_upload(
    db_session: AsyncSession,
    sample_workbook_bytes: bytes,
) -> None:
    result = await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=sample_workbook_bytes,
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    transaction_count = await db_session.scalar(select(func.count()).select_from(Transaction))
    asset_snapshot_count = await db_session.scalar(select(func.count()).select_from(AssetSnapshot))
    investment_count = await db_session.scalar(select(func.count()).select_from(Investment))
    loan_count = await db_session.scalar(select(func.count()).select_from(Loan))
    upload_log = await db_session.scalar(select(UploadLog))

    assert result.tx_total == 2219
    assert result.tx_new == 2219
    assert result.tx_skipped == 0
    assert result.asset_snapshot_count == 45
    assert result.investment_count == 11
    assert result.loan_count == 5
    assert result.status == "success"
    assert transaction_count == 2219
    assert asset_snapshot_count == 45
    assert investment_count == 11
    assert loan_count == 5
    assert upload_log is not None
    assert upload_log.filename == "finance_sample.xlsx"
    assert upload_log.tx_new == 2219
    assert upload_log.status == "success"
    first_transaction = await db_session.scalar(select(Transaction).order_by(Transaction.id.asc()))
    assert first_transaction is not None
    assert first_transaction.merchant == first_transaction.description


async def test_import_transactions_skips_rows_already_loaded(
    db_session: AsyncSession,
    sample_workbook_bytes: bytes,
) -> None:
    await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=sample_workbook_bytes,
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    second = await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=sample_workbook_bytes,
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    transaction_count = await db_session.scalar(select(func.count()).select_from(Transaction))

    assert second.tx_total == 2219
    assert second.tx_new == 0
    assert second.tx_skipped == 2219
    assert transaction_count == 2219


async def test_import_transactions_replaces_preexisting_imported_rows_inside_window(
    db_session: AsyncSession,
    sample_workbook_bytes: bytes,
) -> None:
    seeded_row = parse_transactions_from_bytes(sample_workbook_bytes)[1]
    db_session.add(Transaction(**seeded_row))
    await db_session.commit()

    result = await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=sample_workbook_bytes,
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    matching_rows = await db_session.scalars(
        select(Transaction).where(*transaction_conditions(seeded_row))
    )

    assert result.tx_new == 2218
    assert result.tx_skipped == 1
    assert len(list(matching_rows)) == 1


async def test_import_transactions_records_partial_when_snapshot_sheet_is_missing(
    db_session: AsyncSession,
    sample_workbook_bytes: bytes,
) -> None:
    workbook = load_workbook(BytesIO(sample_workbook_bytes), data_only=True)
    del workbook["뱅샐현황"]
    broken_workbook_bytes = BytesIO()
    workbook.save(broken_workbook_bytes)

    result = await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=broken_workbook_bytes.getvalue(),
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    transaction_count = await db_session.scalar(select(func.count()).select_from(Transaction))
    asset_snapshot_count = await db_session.scalar(select(func.count()).select_from(AssetSnapshot))
    upload_log = await db_session.scalar(select(UploadLog))

    assert result.tx_new == 2219
    assert result.asset_snapshot_count == 0
    assert result.investment_count == 0
    assert result.loan_count == 0
    assert result.status == "partial"
    assert transaction_count == 2219
    assert asset_snapshot_count == 0
    assert upload_log is not None
    assert upload_log.status == "partial"
    assert "뱅샐현황" in (upload_log.error_message or "")


async def test_import_transactions_replaces_snapshot_rows_for_same_snapshot_date(
    db_session: AsyncSession,
    sample_workbook_bytes: bytes,
) -> None:
    first_result = await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=sample_workbook_bytes,
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )
    assert first_result.status == "success"

    workbook = load_workbook(BytesIO(sample_workbook_bytes), data_only=True)
    worksheet = workbook["뱅샐현황"]
    asset_row_index = next(
        index
        for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1)
        if len(row) > 2 and row[2] == "KB국민ONE통장-저축예금"
    )
    worksheet[f"E{asset_row_index}"] = 123456789
    updated_workbook_bytes = BytesIO()
    workbook.save(updated_workbook_bytes)

    second_result = await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=updated_workbook_bytes.getvalue(),
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    asset_snapshot_count = await db_session.scalar(select(func.count()).select_from(AssetSnapshot))
    first_asset_amount = await db_session.scalar(
        select(AssetSnapshot.amount)
        .where(AssetSnapshot.snapshot_date == date(2026, 3, 24))
        .where(AssetSnapshot.side == "asset")
        .where(AssetSnapshot.product_name == "KB국민ONE통장-저축예금")
    )

    assert second_result.tx_new == 0
    assert second_result.asset_snapshot_count == 45
    assert second_result.status == "success"
    assert asset_snapshot_count == 45
    assert first_asset_amount == 123456789


async def test_import_transactions_keeps_existing_rows_and_appends_exact_new_rows_from_rolling_window(
    db_session: AsyncSession,
    sample_workbook_bytes: bytes,
    rolling_window_workbook_bytes: bytes,
) -> None:
    await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=sample_workbook_bytes,
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    result = await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=rolling_window_workbook_bytes,
        filename="sample_260324.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    existing_transactions = list((await db_session.scalars(select(Transaction))).all())
    previous_rows = parse_transactions_from_bytes(sample_workbook_bytes)
    latest_rows = parse_transactions_from_bytes(rolling_window_workbook_bytes)

    assert result.status == "success"
    assert result.tx_total == len(latest_rows)
    assert result.tx_new == 68
    assert result.tx_skipped == 2158
    assert len(existing_transactions) == len(previous_rows) + 68


async def test_import_transactions_does_not_append_duplicate_when_later_window_only_changes_time_or_category(
    db_session: AsyncSession,
    sample_workbook_bytes: bytes,
    rolling_window_workbook_bytes: bytes,
) -> None:
    old_rows = parse_transactions_from_bytes(sample_workbook_bytes)
    new_rows = parse_transactions_from_bytes(rolling_window_workbook_bytes)
    old_row, new_row = find_logically_matching_rows_with_changed_exact_signature(old_rows, new_rows)

    await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=sample_workbook_bytes,
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    existing_row = await db_session.scalar(
        select(Transaction).where(*transaction_conditions(old_row))
    )
    assert existing_row is not None
    existing_row.category_major_user = "사용자수정"
    existing_row.category_minor_user = "세부수정"
    existing_row.memo = "preserve me"
    existing_row.is_deleted = True
    await db_session.commit()

    result = await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=rolling_window_workbook_bytes,
        filename="sample_260324.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    appended_row = await db_session.scalar(
        select(Transaction).where(*transaction_conditions(new_row))
    )
    old_row_after_import = await db_session.scalar(
        select(Transaction).where(*transaction_conditions(old_row))
    )

    assert result.tx_new == 68
    assert appended_row is None
    assert old_row_after_import is not None
    assert old_row_after_import.category_major_user == "사용자수정"
    assert old_row_after_import.category_minor_user == "세부수정"
    assert old_row_after_import.memo == "preserve me"
    assert old_row_after_import.is_deleted is True


def parse_transactions_from_bytes(sample_workbook_bytes: bytes) -> list[dict[str, object]]:
    workbook = load_workbook(BytesIO(sample_workbook_bytes), data_only=True)
    return parse_transactions(workbook)


def transaction_datetime(row: dict[str, object] | Transaction) -> datetime:
    if isinstance(row, Transaction):
        return datetime.combine(row.date, row.time)
    return datetime.combine(row["date"], row["time"])


def transaction_signature(row: dict[str, object] | Transaction) -> tuple[object, ...]:
    if isinstance(row, Transaction):
        return (
            row.date,
            row.time,
            row.type,
            row.category_major,
            row.category_minor,
            row.description,
            row.amount,
            row.currency,
            row.payment_method,
            row.memo,
        )
    return (
        row["date"],
        row["time"],
        row["type"],
        row["category_major"],
        row["category_minor"],
        row["description"],
        row["amount"],
        row["currency"],
        row["payment_method"],
        row["memo"],
    )


def transaction_reconciliation_key(
    row: dict[str, object] | Transaction,
) -> tuple[object, ...]:
    if isinstance(row, Transaction):
        return (
            row.date,
            row.type,
            row.category_major,
            row.category_minor,
            row.description,
            row.amount,
            row.currency,
            row.payment_method,
        )
    return (
        row["date"],
        row["type"],
        row["category_major"],
        row["category_minor"],
        row["description"],
        row["amount"],
        row["currency"],
        row["payment_method"],
    )


def find_logically_matching_rows_with_changed_exact_signature(
    old_rows: list[dict[str, object]],
    new_rows: list[dict[str, object]],
) -> tuple[dict[str, object], dict[str, object]]:
    old_groups: dict[tuple[object, ...], list[dict[str, object]]] = {}
    new_groups: dict[tuple[object, ...], list[dict[str, object]]] = {}

    for row in old_rows:
        old_groups.setdefault(transaction_reconciliation_key(row), []).append(row)
    for row in new_rows:
        new_groups.setdefault(transaction_reconciliation_key(row), []).append(row)

    for key in old_groups.keys() & new_groups.keys():
        old_group = sorted(old_groups[key], key=transaction_datetime)
        new_group = sorted(new_groups[key], key=transaction_datetime)
        for old_row, new_row in zip(old_group, new_group):
            if transaction_signature(old_row) != transaction_signature(new_row):
                return old_row, new_row

    raise AssertionError("Expected at least one logical match with changed exact signature")


def transaction_conditions(row: dict[str, object]) -> tuple[object, ...]:
    category_minor = row["category_minor"]
    payment_method = row["payment_method"]
    return (
        Transaction.date == row["date"],
        Transaction.time == row["time"],
        Transaction.type == row["type"],
        Transaction.category_major == row["category_major"],
        Transaction.category_minor.is_(None)
        if category_minor is None
        else Transaction.category_minor == category_minor,
        Transaction.description == row["description"],
        Transaction.amount == row["amount"],
        Transaction.currency == row["currency"],
        Transaction.payment_method.is_(None)
        if payment_method is None
        else Transaction.payment_method == payment_method,
    )
