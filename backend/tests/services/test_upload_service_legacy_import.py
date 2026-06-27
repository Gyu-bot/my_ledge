from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.asset_snapshot import AssetSnapshot
from app.models.investment import Investment
from app.models.loan import Loan
from app.models.transaction import Transaction
from app.models.upload_log import UploadLog
from app.services.upload_service import import_transactions_from_workbook
from test_upload_service_legacy_helpers import (
    parse_transactions_from_bytes,
    transaction_conditions,
)


async def test_import_transactions_inserts_all_rows_on_first_upload(
    db_session: AsyncSession,
    sample_workbook_bytes: bytes,
) -> None:
    result = await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=sample_workbook_bytes,
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    transaction_count = await db_session.scalar(
        select(func.count()).select_from(Transaction)
    )
    asset_snapshot_count = await db_session.scalar(
        select(func.count()).select_from(AssetSnapshot)
    )
    investment_count = await db_session.scalar(
        select(func.count()).select_from(Investment)
    )
    loan_count = await db_session.scalar(select(func.count()).select_from(Loan))
    upload_log = await db_session.scalar(select(UploadLog))

    assert result.tx_total == 2357
    assert result.tx_new == 2357
    assert result.tx_skipped == 0
    assert result.asset_snapshot_count == 42
    assert result.investment_count == 9
    assert result.loan_count == 4
    assert result.status == "success"
    assert transaction_count == 2357
    assert asset_snapshot_count == 42
    assert investment_count == 9
    assert loan_count == 4
    assert upload_log is not None
    assert upload_log.filename == "finance_sample.xlsx"
    assert upload_log.tx_new == 2357
    assert upload_log.status == "success"
    first_transaction = await db_session.scalar(
        select(Transaction).order_by(Transaction.id.asc())
    )
    assert first_transaction is not None
    assert first_transaction.merchant == first_transaction.description


async def test_import_transactions_retains_latest_five_original_uploads(
    db_session: AsyncSession,
    sample_workbook_bytes: bytes,
    tmp_path: Path,
) -> None:
    upload_dir = tmp_path / "uploads"

    for index in range(6):
        await import_transactions_from_workbook(
            db_session=db_session,
            file_bytes=sample_workbook_bytes,
            filename=f"finance sample {index}.xlsx",
            snapshot_date=date(2026, 3, 24),
            persist_upload_file=True,
            upload_dir=upload_dir,
        )

    saved_uploads = sorted(path.name for path in upload_dir.iterdir())

    assert len(saved_uploads) == 5
    assert saved_uploads == [
        "000002-finance-sample-1.xlsx",
        "000003-finance-sample-2.xlsx",
        "000004-finance-sample-3.xlsx",
        "000005-finance-sample-4.xlsx",
        "000006-finance-sample-5.xlsx",
    ]


async def test_import_transactions_skips_rows_already_loaded(
    db_session: AsyncSession,
    sample_workbook_bytes: bytes,
) -> None:
    await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=sample_workbook_bytes,
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    second = await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=sample_workbook_bytes,
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    transaction_count = await db_session.scalar(
        select(func.count()).select_from(Transaction)
    )

    assert second.tx_total == 2357
    assert second.tx_new == 0
    assert second.tx_skipped == 2357
    assert transaction_count == 2357


async def test_import_transactions_replaces_preexisting_imported_rows_inside_window(
    db_session: AsyncSession,
    sample_workbook_bytes: bytes,
) -> None:
    seeded_row = parse_transactions_from_bytes(sample_workbook_bytes)[1]
    db_session.add(Transaction(**seeded_row))
    await db_session.commit()

    result = await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=sample_workbook_bytes,
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    matching_rows = await db_session.scalars(
        select(Transaction).where(*transaction_conditions(seeded_row))
    )

    assert result.tx_new == 2356
    assert result.tx_skipped == 1
    assert len(list(matching_rows)) == 1


async def test_import_transactions_records_partial_when_snapshot_sheet_is_missing(
    db_session: AsyncSession,
    sample_workbook_bytes: bytes,
) -> None:
    workbook = load_workbook(BytesIO(sample_workbook_bytes), data_only=True)
    del workbook["뱅샐현황"]
    broken_workbook_bytes = BytesIO()
    workbook.save(broken_workbook_bytes)

    result = await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=broken_workbook_bytes.getvalue(),
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    transaction_count = await db_session.scalar(
        select(func.count()).select_from(Transaction)
    )
    asset_snapshot_count = await db_session.scalar(
        select(func.count()).select_from(AssetSnapshot)
    )
    upload_log = await db_session.scalar(select(UploadLog))

    assert result.tx_new == 2357
    assert result.asset_snapshot_count == 0
    assert result.investment_count == 0
    assert result.loan_count == 0
    assert result.status == "partial"
    assert transaction_count == 2357
    assert asset_snapshot_count == 0
    assert upload_log is not None
    assert upload_log.status == "partial"
    assert "뱅샐현황" in (upload_log.error_message or "")


async def test_import_transactions_replaces_snapshot_rows_for_same_snapshot_date(
    db_session: AsyncSession,
    sample_workbook_bytes: bytes,
) -> None:
    first_result = await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=sample_workbook_bytes,
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )
    assert first_result.status == "success"

    existing_asset = await db_session.scalar(
        select(AssetSnapshot)
        .where(AssetSnapshot.snapshot_date == date(2026, 3, 24))
        .where(AssetSnapshot.side == "asset")
        .where(AssetSnapshot.product_name == "KB국민ONE통장-저축예금")
    )
    existing_loan = await db_session.scalar(
        select(Loan)
        .where(Loan.snapshot_date == date(2026, 3, 24))
        .limit(1)
    )
    assert existing_asset is not None
    assert existing_loan is not None
    existing_asset.liquidity_tier = "immediate"
    existing_asset.is_cash_equivalent = True
    existing_loan.monthly_payment = Decimal("650000.00")
    existing_loan.repayment_method = "principal_interest"
    existing_loan_key = (existing_loan.lender, existing_loan.product_name)
    await db_session.commit()

    workbook = load_workbook(BytesIO(sample_workbook_bytes), data_only=True)
    worksheet = workbook["뱅샐현황"]
    asset_row_index = next(
        index
        for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1)
        if len(row) > 2 and row[2] == "KB국민ONE통장-저축예금"
    )
    worksheet[f"E{asset_row_index}"] = 123456789
    updated_workbook_bytes = BytesIO()
    workbook.save(updated_workbook_bytes)

    second_result = await import_transactions_from_workbook(
        db_session=db_session,
        file_bytes=updated_workbook_bytes.getvalue(),
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
    )

    asset_snapshot_count = await db_session.scalar(
        select(func.count()).select_from(AssetSnapshot)
    )
    first_asset_amount = await db_session.scalar(
        select(AssetSnapshot.amount)
        .where(AssetSnapshot.snapshot_date == date(2026, 3, 24))
        .where(AssetSnapshot.side == "asset")
        .where(AssetSnapshot.product_name == "KB국민ONE통장-저축예금")
    )
    replaced_asset = await db_session.scalar(
        select(AssetSnapshot)
        .where(AssetSnapshot.snapshot_date == date(2026, 3, 24))
        .where(AssetSnapshot.side == "asset")
        .where(AssetSnapshot.product_name == "KB국민ONE통장-저축예금")
    )
    replaced_loan = await db_session.scalar(
        select(Loan)
        .where(Loan.snapshot_date == date(2026, 3, 24))
        .where(Loan.lender == existing_loan_key[0])
        .where(Loan.product_name == existing_loan_key[1])
    )

    assert second_result.tx_new == 0
    assert second_result.asset_snapshot_count == 42
    assert second_result.status == "success"
    assert asset_snapshot_count == 42
    assert first_asset_amount == 123456789
    assert replaced_asset is not None
    assert replaced_asset.liquidity_tier == "immediate"
    assert replaced_asset.is_cash_equivalent is True
    assert replaced_loan is not None
    assert replaced_loan.monthly_payment == Decimal("650000.00")
    assert replaced_loan.repayment_method == "principal_interest"
