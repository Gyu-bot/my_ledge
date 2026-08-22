from datetime import date, time
from decimal import Decimal
from io import BytesIO
import json
from pathlib import Path

from openpyxl import Workbook
import pytest
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.asset_snapshot import AssetSnapshot
from app.models.insurance_contract import InsuranceContract
from app.models.installment_plan import InstallmentPlan
from app.models.installment_transaction_link import InstallmentTransactionLink
from app.models.investment import Investment
from app.models.loan import Loan
from app.models.purchase_gate_review import PurchaseGateReview
from app.models.transaction import Transaction
from app.models.transaction import TransactionSourceLifecycleStatus
from app.models.upload_log import UploadLog
from app.models.user_profile_snapshot import UserProfileSnapshot
from app.parsers.snapshots import SnapshotParseResult
from app.parsers.transactions import TransactionRow
from app.schemas.upload import UploadApplyRequest, UploadApplySelection
from app.services import transaction_source_identity
from app.services import upload_apply_service
from app.services.upload_apply_service import (
    apply_transaction_upload_workbook,
    apply_transaction_upload_from_rows,
)
from app.services.upload_preview_service import preview_transaction_upload_from_rows


def build_preview_transaction_row(
    *,
    tx_date: date,
    tx_time: time,
    description: str,
    amount: int,
    category_major: str,
    category_minor: str | None = "기타",
    payment_method: str | None = "체크카드",
    memo: str | None = None,
) -> TransactionRow:
    return {
        "date": tx_date,
        "time": tx_time,
        "type": "지출",
        "category_major": category_major,
        "category_minor": category_minor,
        "description": description,
        "merchant": description,
        "amount": amount,
        "currency": "KRW",
        "payment_method": payment_method,
        "memo": memo,
    }


def build_preview_transaction(
    *,
    tx_date: date,
    tx_time: time,
    description: str,
    amount: int,
    category_major: str,
    category_minor: str | None = "기타",
    payment_method: str | None = "체크카드",
) -> Transaction:
    return Transaction(
        date=tx_date,
        time=tx_time,
        type="지출",
        category_major=category_major,
        category_minor=category_minor,
        description=description,
        merchant=description,
        amount=amount,
        currency="KRW",
        payment_method=payment_method,
        memo=None,
        source="import",
    )


def build_fixture_workbook_bytes() -> bytes:
    workbook = Workbook()
    workbook.active.title = "fixture"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_snapshot_fixture(*, asset_amount: str) -> SnapshotParseResult:
    return SnapshotParseResult(
        asset_snapshots=[
            {
                "side": "asset",
                "category": "fixture-category",
                "product_name": "fixture-asset",
                "amount": Decimal(asset_amount),
            }
        ],
        insurance_contracts=[
            {
                "insurer": "fixture-insurer",
                "product_name": "fixture-insurance",
                "contract_status": "active",
                "total_paid": Decimal("10"),
                "contract_date": None,
                "maturity_date": None,
            }
        ],
        investments=[
            {
                "product_type": "fund",
                "broker": "fixture-broker",
                "product_name": "fixture-investment",
                "cost_basis": Decimal("20"),
                "market_value": Decimal("21"),
                "return_rate": Decimal("5"),
            }
        ],
        loans=[
            {
                "loan_type": "credit",
                "lender": "fixture-lender",
                "product_name": "fixture-loan",
                "principal": Decimal("30"),
                "balance": Decimal("25"),
                "interest_rate": Decimal("3"),
                "start_date": None,
                "maturity_date": None,
            },
            {
                "loan_type": "credit",
                "lender": "fixture-lender",
                "product_name": "fixture-loan",
                "principal": Decimal("40"),
                "balance": Decimal("35"),
                "interest_rate": Decimal("4"),
                "start_date": None,
                "maturity_date": None,
            },
        ],
        user_profile={"gender": "fixture", "age": 30, "credit_score_kcb": 800},
    )


async def build_apply_request(
    db_session: AsyncSession,
    rows: list[TransactionRow],
) -> UploadApplyRequest:
    preview = await preview_transaction_upload_from_rows(db_session, rows)
    change = preview.safe_changes[0]
    return UploadApplyRequest(
        confirmation=True,
        selections=[
            UploadApplySelection(
                change_type=change.change_type,
                source_row_hash=change.source_row_hash or "",
                existing_transaction_id=change.existing_transaction_id,
            )
        ],
    )


async def test_apply_workbook_persists_and_replaces_snapshots_and_original(
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    tmp_path: Path,
) -> None:
    workbook_bytes = build_fixture_workbook_bytes()
    rows = [
        build_preview_transaction_row(
            tx_date=date(2026, 3, 24),
            tx_time=time(9, 0),
            description="fixture-transaction",
            amount=-100,
            category_major="fixture-category",
        )
    ]
    snapshots = build_snapshot_fixture(asset_amount="100")
    estimate_keys: list[list[tuple[str, str]]] = []

    async def capture_estimates(
        _db_session: AsyncSession,
        *,
        loan_keys: list[tuple[str, str]],
    ) -> None:
        estimate_keys.append(loan_keys)

    monkeypatch.setattr(
        upload_apply_service,
        "parse_upload_workbook_contents",
        lambda **_: (rows, snapshots),
    )
    monkeypatch.setattr(
        upload_apply_service,
        "apply_loan_repayment_estimates_for_latest_snapshots",
        capture_estimates,
    )

    upload_dir = tmp_path / "uploads"
    upload_dir.mkdir()
    for index in range(5):
        (upload_dir / f"00000{index}-old.xlsx").write_bytes(b"old")

    first = await apply_transaction_upload_workbook(
        db_session=db_session,
        file_bytes=workbook_bytes,
        filename="../../unsafe name.xlsx",
        snapshot_date=date(2026, 3, 24),
        apply_request=await build_apply_request(db_session, rows),
        upload_dir=upload_dir,
    )

    asset = await db_session.scalar(select(AssetSnapshot))
    loan = await db_session.scalar(
        select(Loan).where(Loan.product_name == "fixture-loan")
    )
    assert asset is not None
    assert loan is not None
    asset.liquidity_tier = "immediate"
    asset.is_cash_equivalent = True
    loan.monthly_payment = Decimal("7")
    loan.monthly_payment_source = "manual"
    loan.repayment_method = "principal_interest"
    loan.repayment_method_source = "manual"
    await db_session.commit()

    snapshots = build_snapshot_fixture(asset_amount="200")
    second = await apply_transaction_upload_workbook(
        db_session=db_session,
        file_bytes=workbook_bytes,
        filename="second.xlsx",
        snapshot_date=date(2026, 3, 24),
        apply_request=await build_apply_request(db_session, rows),
        upload_dir=upload_dir,
    )

    replaced_asset = await db_session.scalar(select(AssetSnapshot))
    replaced_loan = await db_session.scalar(
        select(Loan).where(Loan.product_name == "fixture-loan")
    )
    assert first.asset_snapshot_count == 1
    assert first.insurance_contract_count == 1
    assert first.investment_count == 1
    assert first.loan_count == 2
    assert second.asset_snapshot_count == 1
    assert await db_session.scalar(select(func.count()).select_from(Transaction)) == 1
    assert await db_session.scalar(select(func.count()).select_from(AssetSnapshot)) == 1
    assert (
        await db_session.scalar(select(func.count()).select_from(InsuranceContract))
        == 1
    )
    assert await db_session.scalar(select(func.count()).select_from(Investment)) == 1
    assert await db_session.scalar(select(func.count()).select_from(Loan)) == 2
    assert (
        await db_session.scalar(select(func.count()).select_from(UserProfileSnapshot))
        == 1
    )
    assert replaced_asset is not None
    assert replaced_asset.amount == Decimal("200")
    assert replaced_asset.liquidity_tier == "immediate"
    assert replaced_asset.is_cash_equivalent is True
    assert replaced_loan is not None
    assert replaced_loan.monthly_payment == Decimal("7")
    assert replaced_loan.monthly_payment_source == "manual"
    assert replaced_loan.repayment_method == "principal_interest"
    assert replaced_loan.repayment_method_source == "manual"
    assert estimate_keys == [
        [("fixture-lender", "fixture-loan"), ("fixture-lender", "fixture-loan (2)")],
        [("fixture-lender", "fixture-loan"), ("fixture-lender", "fixture-loan (2)")],
    ]
    assert (upload_dir / "000001-unsafe-name.xlsx").read_bytes() == workbook_bytes
    assert (upload_dir / "000002-second.xlsx").read_bytes() == workbook_bytes
    assert not (upload_dir / "000000-old.xlsx").exists()
    assert len(list(upload_dir.iterdir())) == 5


async def test_apply_workbook_records_nonfatal_warning_when_original_upload_save_fails(
    monkeypatch: pytest.MonkeyPatch,
    db_session: AsyncSession,
    tmp_path: Path,
) -> None:
    workbook_bytes = build_fixture_workbook_bytes()
    rows = [
        build_preview_transaction_row(
            tx_date=date(2026, 3, 24),
            tx_time=time(9, 0),
            description="fixture-transaction",
            amount=-100,
            category_major="fixture-category",
        )
    ]
    snapshots = build_snapshot_fixture(asset_amount="100")

    monkeypatch.setattr(
        upload_apply_service,
        "parse_upload_workbook_contents",
        lambda **_: (rows, snapshots),
    )

    def fail_persist(**_: object) -> None:
        raise OSError("disk full")

    monkeypatch.setattr(upload_apply_service, "persist_original_upload", fail_persist)

    result = await apply_transaction_upload_workbook(
        db_session=db_session,
        file_bytes=workbook_bytes,
        filename="warning.xlsx",
        snapshot_date=date(2026, 3, 24),
        apply_request=await build_apply_request(db_session, rows),
        upload_dir=tmp_path / "uploads",
    )

    upload_log = await db_session.scalar(
        select(UploadLog).where(UploadLog.id == result.upload_id)
    )

    assert result.upload_id == 1
    assert await db_session.scalar(select(func.count()).select_from(Transaction)) == 1
    assert upload_log is not None
    assert upload_log.status == "success"
    assert "warning: original upload not saved: disk full" in (
        upload_log.error_message or ""
    )


async def test_apply_new_rows_creates_transactions_and_upload_log(
    db_session: AsyncSession,
) -> None:
    incoming = [
        build_preview_transaction_row(
            tx_date=date(2026, 3, 24),
            tx_time=time(9, 0),
            description="커피",
            amount=-4500,
            category_major="식비",
        )
    ]

    result = await apply_transaction_upload_from_rows(
        db_session=db_session,
        parsed_rows=incoming,
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
        apply_request=UploadApplyRequest(
            confirmation=True,
            selections=[
                UploadApplySelection(
                    change_type="new",
                    source_row_hash=transaction_source_identity.source_row_hash_from_row(
                        incoming[0]
                    ),
                    existing_transaction_id=None,
                )
            ],
        ),
    )

    transaction = await db_session.scalar(select(Transaction))
    upload_log = await db_session.scalar(
        select(UploadLog).where(UploadLog.id == result.upload_id)
    )

    assert result.tx_new == 1
    assert result.selected_change_count == 1
    assert result.applied_change_count == 1
    assert transaction is not None
    assert transaction.description == "커피"
    assert (
        transaction.source_lifecycle_status
        == TransactionSourceLifecycleStatus.ACTIVE.value
    )
    assert upload_log is not None
    assert upload_log.reconciliation_mode == "explicit_apply"


async def test_apply_source_fields_changed_preserves_user_overrides(
    db_session: AsyncSession,
) -> None:
    existing = build_preview_transaction(
        tx_date=date(2026, 3, 25),
        tx_time=time(12, 30),
        description="정류장",
        amount=-1500,
        category_major="교통",
        category_minor="대중교통",
    )
    existing.category_major_user = "사용자카테고리"
    existing.memo = "기록"
    existing.merchant = "사용자상점"
    db_session.add(existing)
    await db_session.commit()

    incoming = [
        build_preview_transaction_row(
            tx_date=date(2026, 3, 25),
            tx_time=time(12, 30),
            description="정류장",
            amount=-1500,
            category_major="생활",
            category_minor="교통",
            payment_method="체크카드",
        )
    ]

    preview = await preview_transaction_upload_from_rows(db_session, incoming)
    assert len(preview.safe_changes) == 1

    source_change = preview.safe_changes[0]
    result = await apply_transaction_upload_from_rows(
        db_session=db_session,
        parsed_rows=incoming,
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
        apply_request=UploadApplyRequest(
            confirmation=True,
            selections=[
                UploadApplySelection(
                    change_type=source_change.change_type,
                    source_row_hash=source_change.source_row_hash or "",
                    existing_transaction_id=existing.id,
                )
            ],
        ),
    )

    stored = await db_session.scalar(
        select(Transaction).where(Transaction.id == existing.id)
    )

    assert result.tx_new == 0
    assert result.applied_change_count == 1
    assert stored is not None
    assert stored.category_major == "생활"
    assert stored.category_major_user == "사용자카테고리"
    assert stored.memo == "기록"
    assert stored.merchant == "사용자상점"


async def test_apply_missing_from_latest_export_marks_lifecycle_status(
    db_session: AsyncSession,
) -> None:
    existing = build_preview_transaction(
        tx_date=date(2026, 3, 26),
        tx_time=time(13, 0),
        description="구독",
        amount=-50000,
        category_major="구독",
    )
    db_session.add(existing)
    await db_session.commit()

    incoming = [
        build_preview_transaction_row(
            tx_date=date(2026, 3, 26),
            tx_time=time(14, 0),
            description="새구독",
            amount=-60000,
            category_major="기타",
            category_minor=None,
            payment_method="체크카드",
        )
    ]

    preview = await preview_transaction_upload_from_rows(db_session, incoming)
    missing_change = next(
        change
        for change in preview.safe_changes
        if change.change_type == "missing_from_latest_export"
    )

    result = await apply_transaction_upload_from_rows(
        db_session=db_session,
        parsed_rows=incoming,
        filename="finance_sample.xlsx",
        snapshot_date=date(2026, 3, 24),
        apply_request=UploadApplyRequest(
            confirmation=True,
            selections=[
                UploadApplySelection(
                    change_type=missing_change.change_type,
                    source_row_hash=missing_change.source_row_hash or "",
                    existing_transaction_id=missing_change.existing_transaction_id,
                )
            ],
        ),
    )

    updated = await db_session.scalar(
        select(Transaction).where(Transaction.id == existing.id)
    )

    assert result.applied_change_count == 1
    assert updated is not None
    assert (
        updated.source_lifecycle_status
        == TransactionSourceLifecycleStatus.MISSING_FROM_LATEST_EXPORT.value
    )


async def test_possible_replacement_requires_review_and_explicit_apply_supersedes_existing(
    db_session: AsyncSession,
) -> None:
    existing = build_preview_transaction(
        tx_date=date(2026, 3, 29),
        tx_time=time(11, 0),
        description="기존 결제",
        amount=-120000,
        category_major="쇼핑",
        category_minor="가전",
    )
    existing.category_major_user = "사용자대분류"
    existing.category_minor_user = "사용자소분류"
    existing.memo = "사용자 메모"
    existing.merchant = "사용자 거래처"
    existing.spend_necessity = "essential"
    existing.recurring_payment_kind = "installment"
    db_session.add(existing)
    await db_session.flush()

    plan = InstallmentPlan(
        display_name="테스트 할부",
        merchant="테스트상점",
        payment_method="체크카드",
        total_installments=3,
        monthly_amount=40000,
        first_payment_date=date(2026, 3, 1),
    )
    db_session.add(plan)
    await db_session.flush()
    db_session.add(
        InstallmentTransactionLink(
            transaction_id=existing.id,
            installment_plan_id=plan.id,
            installment_number=2,
            source="manual",
            memo="연결 유지",
        )
    )
    db_session.add(
        PurchaseGateReview(
            candidate_key=f"transaction:{existing.id}",
            candidate_type="transaction",
            transaction_id=existing.id,
            review_status="reviewed",
            memo="검토 유지",
        )
    )
    await db_session.commit()

    incoming = [
        build_preview_transaction_row(
            tx_date=date(2026, 3, 29),
            tx_time=time(14, 30),
            description="교체 결제",
            amount=-120000,
            category_major="생활",
            category_minor="리빙",
            payment_method="체크카드",
        )
    ]

    preview = await preview_transaction_upload_from_rows(db_session, incoming)

    assert preview.safe_change_count == 0
    assert preview.review_required_count == 1
    review_change = preview.review_required_changes[0]
    assert review_change.change_type == "possible_replacement"
    assert review_change.review_required is True
    assert review_change.auto_apply_safe is False
    assert review_change.existing_transaction_id == existing.id

    result = await apply_transaction_upload_from_rows(
        db_session=db_session,
        parsed_rows=incoming,
        filename="possible_replacement_apply.xlsx",
        snapshot_date=date(2026, 3, 29),
        apply_request=UploadApplyRequest(
            confirmation=True,
            selections=[
                UploadApplySelection(
                    change_type=review_change.change_type,
                    source_row_hash=review_change.source_row_hash or "",
                    existing_transaction_id=review_change.existing_transaction_id,
                )
            ],
        ),
    )

    rows = list(
        (
            await db_session.scalars(select(Transaction).order_by(Transaction.id.asc()))
        ).all()
    )
    upload_log = await db_session.scalar(
        select(UploadLog).where(UploadLog.id == result.upload_id)
    )
    existing_after = next(row for row in rows if row.id == existing.id)
    replacement = next(row for row in rows if row.id != existing.id)
    moved_installment_link = await db_session.scalar(
        select(InstallmentTransactionLink).where(
            InstallmentTransactionLink.transaction_id == replacement.id
        )
    )
    old_installment_link = await db_session.scalar(
        select(InstallmentTransactionLink).where(
            InstallmentTransactionLink.transaction_id == existing.id
        )
    )
    review = await db_session.scalar(select(PurchaseGateReview))

    assert result.tx_new == 1
    assert result.tx_skipped == 0
    assert result.applied_change_count == 1
    assert result.applied_changes[0].change_type == "possible_replacement"
    assert len(rows) == 2
    assert (
        existing_after.source_lifecycle_status
        == TransactionSourceLifecycleStatus.SUPERSEDED.value
    )
    assert existing_after.superseded_by_transaction_id == replacement.id
    assert existing_after.description == "기존 결제"
    assert (
        replacement.source_lifecycle_status
        == TransactionSourceLifecycleStatus.ACTIVE.value
    )
    assert replacement.description == "교체 결제"
    assert replacement.category_major == "생활"
    assert replacement.category_minor == "리빙"
    assert replacement.category_major_user == "사용자대분류"
    assert replacement.category_minor_user == "사용자소분류"
    assert replacement.memo == "사용자 메모"
    assert replacement.merchant == "사용자 거래처"
    assert replacement.spend_necessity == "essential"
    assert replacement.recurring_payment_kind == "installment"
    assert replacement.first_seen_import_id == result.upload_id
    assert replacement.last_seen_import_id == result.upload_id
    assert replacement.source_row_hash is not None
    assert old_installment_link is None
    assert moved_installment_link is not None
    assert moved_installment_link.installment_plan_id == plan.id
    assert moved_installment_link.installment_number == 2
    assert moved_installment_link.memo == "연결 유지"
    assert review is not None
    assert review.transaction_id == replacement.id
    assert review.candidate_key == f"transaction:{replacement.id}"
    assert review.review_status == "reviewed"
    assert review.memo == "검토 유지"
    assert upload_log is not None
    assert upload_log.tx_new == 1
    audit_payload = json.loads(upload_log.reconciliation_audit or "{}")
    assert audit_payload["applied_change_count"] == 1
    assert audit_payload["change_type_counts"]["possible_replacement"] == 1
